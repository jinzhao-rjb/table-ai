#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
表格生成与测试脚本
1. 生成各种类型的表格
2. 使用计算方法测试表格，不包含具体函数实现
"""

import os
import sys
import random
import json
from datetime import datetime
from openpyxl import Workbook

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.abspath(__file__))

# 创建输出目录
table_output_dir = os.path.join(project_root, "generated_tables")
os.makedirs(table_output_dir, exist_ok=True)
test_results_dir = os.path.join(project_root, "table_test_results")
os.makedirs(test_results_dir, exist_ok=True)

def generate_table_data(rows=10, cols=5):
    """
    生成表格数据
    """
    data = []
    
    # 生成表头
    headers = [f"列{i+1}" for i in range(cols)]
    data.append(headers)
    
    # 生成数据行
    for _ in range(rows):
        row = [random.randint(1, 100) for _ in range(cols)]
        data.append(row)
    
    return data

def save_table_to_excel(data, filename):
    """
    将表格数据保存为Excel文件
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    
    # 写入数据
    for row in data:
        ws.append(row)
    
    # 保存文件
    filepath = os.path.join(table_output_dir, filename.replace('.json', '.xlsx'))
    wb.save(filepath)
    print(f"已保存表格: {os.path.basename(filepath)}")
    return filepath

def read_table_from_excel(filepath):
    """
    从Excel文件读取表格数据
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    return data

def calculate_table_stats(table_data):
    """
    计算表格统计信息
    """
    if not table_data or len(table_data) < 2:
        return {"error": "表格数据为空或只有表头"}
    
    headers = table_data[0]
    rows = table_data[1:]
    cols = len(headers)
    total_rows = len(rows)
    
    # 计算每列的统计信息
    stats = {
        "total_rows": total_rows,
        "total_cols": cols,
        "columns": {}
    }
    
    for col_idx, header in enumerate(headers):
        # 提取列数据
        col_data = [row[col_idx] for row in rows]
        
        # 计算统计值
        col_sum = sum(col_data)
        col_avg = col_sum / len(col_data)
        col_min = min(col_data)
        col_max = max(col_data)
        col_median = sorted(col_data)[len(col_data) // 2]
        
        stats["columns"][header] = {
            "sum": col_sum,
            "average": col_avg,
            "min": col_min,
            "max": col_max,
            "median": col_median
        }
    
    return stats

def calculate_row_stats(table_data):
    """
    计算行统计信息
    """
    if not table_data or len(table_data) < 2:
        return {"error": "表格数据为空或只有表头"}
    
    headers = table_data[0]
    rows = table_data[1:]
    
    # 计算每行的统计信息
    row_stats = []
    for i, row in enumerate(rows, 1):
        row_sum = sum(row)
        row_avg = row_sum / len(row)
        row_min = min(row)
        row_max = max(row)
        
        row_stats.append({
            "row": i,
            "sum": row_sum,
            "average": row_avg,
            "min": row_min,
            "max": row_max
        })
    
    return row_stats

def calculate_table_total(table_data):
    """
    计算表格总和
    """
    if not table_data or len(table_data) < 2:
        return 0
    
    rows = table_data[1:]
    total = 0
    for row in rows:
        total += sum(row)
    
    return total

def generate_and_save_tables():
    """
    生成各种类型的表格并保存
    包括10种数据库已有的常见表格类型和10种需要学习的复杂表格类型
    """
    print("开始生成表格...")
    
    # 生成10种数据库已有的常见表格类型
    common_table_types = [
        {"name": "simple_table", "rows": 10, "cols": 5, "type": "common"},
        {"name": "small_table", "rows": 5, "cols": 3, "type": "common"},
        {"name": "large_table", "rows": 100, "cols": 10, "type": "common"},
        {"name": "wide_table", "rows": 10, "cols": 20, "type": "common"},
        {"name": "tall_table", "rows": 200, "cols": 5, "type": "common"},
        {"name": "medium_table", "rows": 50, "cols": 7, "type": "common"},
        {"name": "minimal_table", "rows": 3, "cols": 2, "type": "common"},
        {"name": "square_table", "rows": 15, "cols": 15, "type": "common"},
        {"name": "long_table", "rows": 300, "cols": 4, "type": "common"},
        {"name": "standard_table", "rows": 25, "cols": 8, "type": "common"}
    ]
    
    # 生成10种需要学习的复杂表格类型
    complex_table_types = [
        {"name": "nested_table_1", "rows": 20, "cols": 10, "type": "complex"},
        {"name": "nested_table_2", "rows": 30, "cols": 15, "type": "complex"},
        {"name": "multi_header_table", "rows": 12, "cols": 8, "type": "complex"},
        {"name": "merged_cells_table", "rows": 15, "cols": 10, "type": "complex"},
        {"name": "irregular_table", "rows": 25, "cols": 12, "type": "complex"},
        {"name": "percentage_table", "rows": 18, "cols": 9, "type": "complex"},
        {"name": "currency_table", "rows": 22, "cols": 7, "type": "complex"},
        {"name": "date_range_table", "rows": 35, "cols": 6, "type": "complex"},
        {"name": "formula_table", "rows": 40, "cols": 8, "type": "complex"},
        {"name": "mixed_data_table", "rows": 50, "cols": 10, "type": "complex"}
    ]
    
    # 合并所有表格类型
    table_types = common_table_types + complex_table_types
    
    saved_files = []
    
    for table_type in table_types:
        # 生成表格数据
        data = generate_table_data(table_type["rows"], table_type["cols"])
        
        # 保存为Excel文件
        filename = f"{table_type['type']}_{table_type['name']}.xlsx"
        filepath = save_table_to_excel(data, filename)
        saved_files.append(filepath)
    
    print(f"\n已生成{len(saved_files)}个表格，保存目录: {table_output_dir}")
    print(f"其中包含10种常见表格类型和10种复杂表格类型")
    return saved_files

def test_tables(table_files):
    """
    测试表格，计算统计信息
    """
    print("\n开始测试表格...")
    
    test_results = {}
    
    for filepath in table_files:
        filename = os.path.basename(filepath)
        print(f"测试表格: {filename}")
        
        # 读取表格数据
        with open(filepath, 'r', encoding='utf-8') as f:
            table_data = json.load(f)
        
        # 计算统计信息
        col_stats = calculate_table_stats(table_data)
        row_stats = calculate_row_stats(table_data)
        total = calculate_table_total(table_data)
        
        # 保存测试结果
        test_results[filename] = {
            "column_stats": col_stats,
            "row_stats": row_stats,
            "total": total,
            "test_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    # 保存所有测试结果
    results_file = os.path.join(test_results_dir, "table_test_results.json")
    with open(results_file, 'w', encoding='utf-8') as f:
        json.dump(test_results, f, ensure_ascii=False, indent=2)
    
    # 保存简洁的测试结果
    simple_results = {}
    for filename, result in test_results.items():
        simple_results[filename] = {
            "total": result["total"],
            "columns": len(result["column_stats"].get("columns", {})),
            "rows": result["column_stats"].get("total_rows", 0)
        }
    
    simple_results_file = os.path.join(test_results_dir, "table_test_summary.txt")
    with open(simple_results_file, 'w', encoding='utf-8') as f:
        f.write(f"表格测试结果\n")
        f.write(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"="*60 + "\n\n")
        
        for filename, result in simple_results.items():
            f.write(f"表格: {filename}\n")
            f.write(f"  行数: {result['rows']}\n")
            f.write(f"  列数: {result['columns']}\n")
            f.write(f"  总和: {result['total']}\n")
            f.write(f"-"*40 + "\n")
    
    print(f"\n测试完成！")
    print(f"详细结果: {results_file}")
    print(f"简洁结果: {simple_results_file}")
    
    return test_results

def main():
    """
    主函数
    """
    print("表格生成与测试系统")
    print("="*50)
    
    # 1. 生成表格
    table_files = generate_and_save_tables()
    
    # 2. 测试表格
    test_results = test_tables(table_files)
    
    print("\n" + "="*50)
    print("所有操作完成！")

if __name__ == "__main__":
    main()
