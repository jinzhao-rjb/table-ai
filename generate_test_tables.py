#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
表格生成脚本
用于生成各种类型的表格，保存在项目中，用于测试表格提取功能
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from faker import Faker

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.abspath(__file__))

# 创建输出目录
table_output_dir = os.path.join(project_root, "test_tables")
os.makedirs(table_output_dir, exist_ok=True)

# 初始化Faker用于生成模拟数据
faker = Faker("zh_CN")

class TableGenerator:
    """
    表格生成器类，用于生成各种类型的表格
    """
    
    def __init__(self):
        """
        初始化表格生成器
        """
        self.faker = Faker("zh_CN")
    
    def generate_simple_table(self, filename="simple_table.xlsx"):
        """
        生成简单表格：基本的行列表格，无合并单元格
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "简单表格"
        
        # 添加表头
        headers = ["ID", "姓名", "年龄", "邮箱", "地址"]
        ws.append(headers)
        
        # 添加数据
        for i in range(1, 21):
            row = [
                i,
                self.faker.name(),
                self.faker.random_int(18, 60),
                self.faker.email(),
                self.faker.address()
            ]
            ws.append(row)
        
        # 保存表格
        wb.save(os.path.join(table_output_dir, filename))
        print(f"已生成简单表格: {filename}")
    
    def generate_merged_cells_table(self, filename="merged_cells_table.xlsx"):
        """
        生成合并单元格表格：包含colspan和rowspan的表格
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "合并单元格表格"
        
        # 添加合并单元格表头
        ws.merge_cells("A1:E1")
        ws["A1"] = "学生信息表"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(["年级", "班级", "姓名", "科目", "成绩"])
        
        # 添加数据，包含