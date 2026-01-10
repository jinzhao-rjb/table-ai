import os 
import cv2 
import numpy as np 
import logging 
from bs4 import BeautifulSoup 
from openpyxl import Workbook 
from openpyxl.styles import Alignment, Border, Side 
from openpyxl.utils import get_column_letter 

logger = logging.getLogger(__name__) 

class TableProcessor: 
    def __init__(self, yolo_model_path): 
        from ultralytics import YOLO 
        # 修复属性报错：加载后立即执行 fuse() 
        self.yolo_model = None
        try: 
            if yolo_model_path is not None:
                self.yolo_model = YOLO(yolo_model_path) 
                self.yolo_model.fuse() 
        except Exception as e: 
            logging.error(f"模型加载异常: {e}") 
        self.device = 'cuda' if cv2.cuda.getCudaEnabledDeviceCount() > 0 else 'cpu' 

    def enhance_image(self, crop_img):
        """增强图像对比度，使边框更明显"""
        # 增加对比度，使浅色边框更明显
        lab = cv2.cvtColor(crop_img, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        cl = clahe.apply(l)
        limg = cv2.merge((cl, a, b))
        enhanced_img = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)
        
        # 优化：检查图片尺寸，过长边超过2048px时resize到1536px
        max_size = 1536
        h, w = enhanced_img.shape[:2]
        if max(h, w) > max_size:
            scale = max_size / max(h, w)
            new_w = int(w * scale)
            new_h = int(h * scale)
            # 使用INTER_AREA插值法，适合缩小图像
            enhanced_img = cv2.resize(enhanced_img, (new_w, new_h), interpolation=cv2.INTER_AREA)
        
        return enhanced_img
    
    def process_image(self, img_path): 
        """使用 YOLO 裁剪表格核心区域，电子版也必须执行此步以去除页码/白边""" 
        conf_threshold = 0.7  # 提高门槛，只处理非常确定的表格
        results = self.yolo_model(img_path)
        confidence = 0.0  # 默认置信度为0
        
        if len(results) > 0 and results[0].boxes:
            # 拿到置信度最高的框 
            max_idx = results[0].boxes.conf.argmax().cpu().item()
            confidence = results[0].boxes.conf[max_idx].cpu().item()
            
            if confidence >= conf_threshold:
                box = results[0].boxes[max_idx].xyxy[0].cpu().numpy()
                img = cv2.imread(img_path)
                x1, y1, x2, y2 = map(int, box)
                w, h = img.shape[1], img.shape[0]
                
                # 修改 padding 为 0 或 2（外扩），绝不内缩
                padding = 0  # 或者设为 2，如果你觉得切得太死，给它留点白边
                
                x1 = max(0 , x1 - padding)
                y1 = max(0 , y1 - padding)
                x2 = min (w, x2 + padding)
                y2 = min (h, y2 + padding)
                
                # 增加你提到的置信度记录逻辑（用于未来的 LoRA 微调）
                if confidence < 0.9:
                    # 记录难例路径，但不影响正常处理
                    logger.info(f"低置信度警告({confidence:.2f}): 建议将图片 {img_path} 加入 LoRA 微调标注集")
                
                crop_img = img[y1:y2, x1:x2]
                
                # 增强图像对比度，使边框更明显
                enhanced_img = self.enhance_image(crop_img)
                return enhanced_img, confidence
                
        # 如果没有检测到表格，也对原图进行增强
        img = cv2.imread(img_path)
        enhanced_img = self.enhance_image(img)
        return enhanced_img, confidence 

    def save_html_to_excel(self, html_str, output_path): 
        """ 
        全动态还原引擎：没有行数限制，支持任意长度 
        """ 
        # 预处理：去掉可能存在的 Markdown 标签，防止 BeautifulSoup 解析失败 
        clean_html = html_str.replace("```html", "").replace("```", "").strip() 
        # 使用lxml解析器，比html.parser更快 
        soup = BeautifulSoup(clean_html, 'lxml') 
        table = soup.find('table') 
        if not table: 
            return False 

        wb = Workbook() 
        ws = wb.active 
        occupied = set()  # 追踪已被占用的单元格坐标 (row, col) 

        # 定义边框样式 
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin')) 

        for r_idx, tr in enumerate(table.find_all('tr', recursive=False), 1): 
            c_idx = 1 
            for td in tr.find_all(['td', 'th']): 
                # 关键：跳过所有已被上方或左侧合并单元格占用的位置 
                while (r_idx, c_idx) in occupied: 
                    c_idx += 1 
                
                # 获取合并属性 
                rs = int(td.get('rowspan', 1)) 
                cs = int(td.get('colspan', 1)) 
                val = td.get_text(strip=True) 

                # 写入起始格子 
                cell = ws.cell(row=r_idx, column=c_idx, value=val) 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
                cell.border = thin_border 

                # 处理物理合并逻辑 
                if rs > 1 or cs > 1: 
                    ws.merge_cells(start_row=r_idx, start_column=c_idx, 
                                   end_row=r_idx + rs - 1, end_column=c_idx + cs - 1) 
                    # 将合并覆盖的范围全部标记为占用 
                    for r in range(r_idx, r_idx + rs): 
                        for c in range(c_idx, c_idx + cs): 
                            occupied.add((r, c)) 
                else: 
                    occupied.add((r_idx, c_idx)) 
                
                c_idx += 1 

        self._auto_adjust_column_width(ws) 
        wb.save(output_path) 
        return True 

    def _auto_adjust_column_width(self, ws): 
        """根据内容自动调整列宽""" 
        for col_idx in range(1, ws.max_column + 1): 
            max_length = 0 
            column = get_column_letter(col_idx) 
            for row in range(1, ws.max_row + 1): 
                cell = ws.cell(row=row, column=col_idx) 
                if cell.value: 
                    length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in str(cell.value)) 
                    max_length = max(max_length, length) 
            ws.column_dimensions[column].width = min(max_length + 2, 50) 
