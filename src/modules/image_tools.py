import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageEnhance
import cv2
import numpy as np
import qrcode
from PIL import ImageOps
# 不再使用本地Tesseract OCR，改用AI OCR

logger = logging.getLogger(__name__)


class ImageTools:
    """图片处理类，提供各种图片操作功能"""

    @staticmethod
    def batch_resize_images(
        files: List[str],
        width: Optional[int] = None,
        height: Optional[int] = None,
        output_dir: Optional[str] = None,
        keep_aspect_ratio: bool = True,
        quality: int = 85
    ) -> List[Dict[str, Any]]:
        """批量调整图片大小

        Args:
            files: 图片文件列表
            width: 目标宽度，None表示保持原比例
            height: 目标高度，None表示保持原比例
            output_dir: 输出目录，None表示原目录
            keep_aspect_ratio: 是否保持宽高比
            quality: 输出图片质量，1-100

        Returns:
            调整结果列表，每个元素包含原路径、新路径和成功状态
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(
                        output_dir) / f"{file_path.stem}_resized{file_path.suffix}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_name(
                        f"{file_path.stem}_resized{file_path.suffix}")

                # 打开图片
                img = Image.open(file_path)

                # 计算新尺寸
                if keep_aspect_ratio:
                    if width and height:
                        # 计算缩放比例
                        img_width, img_height = img.size
                        ratio = min(width / img_width, height / img_height)
                        new_width = int(img_width * ratio)
                        new_height = int(img_height * ratio)
                    elif width:
                        # 仅指定宽度，按比例计算高度
                        img_width, img_height = img.size
                        ratio = width / img_width
                        new_width = width
                        new_height = int(img_height * ratio)
                    elif height:
                        # 仅指定高度，按比例计算宽度
                        img_width, img_height = img.size
                        ratio = height / img_height
                        new_width = int(img_width * ratio)
                        new_height = height
                    else:
                        # 不调整大小
                        new_width, new_height = img.size
                else:
                    # 不保持比例
                    new_width = width or img.width
                    new_height = height or img.height

                # 调整大小
                resized_img = img.resize(
                    (new_width, new_height), Image.LANCZOS)

                # 保存图片
                resized_img.save(output_path, quality=quality)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True,
                    "old_size": img.size,
                    "new_size": (new_width, new_height)
                })
                logger.info(f"图片调整大小成功: {file_path} -> {output_path}")
            except Exception as e:
                logger.error(f"图片调整大小失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def batch_convert_format(
        files: List[str],
        output_format: str,
        output_dir: Optional[str] = None,
        quality: int = 85
    ) -> List[Dict[str, Any]]:
        """批量转换图片格式

        Args:
            files: 图片文件列表
            output_format: 输出格式，如png, jpg, webp
            output_dir: 输出目录，None表示原目录
            quality: 输出图片质量，1-100

        Returns:
            转换结果列表
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(output_dir) / \
                        f"{file_path.stem}.{output_format.lower()}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_suffix(
                        f".{output_format.lower()}")

                # 打开图片
                img = Image.open(file_path)

                # 转换格式
                if output_format.lower() == "jpg" or output_format.lower() == "jpeg":
                    # 处理透明背景
                    if img.mode in ("RGBA", "LA"):
                        background = Image.new(
                            "RGB", img.size, (255, 255, 255))
                        background.paste(img, mask=img.split()[3])
                        img = background
                    img.save(output_path, quality=quality)
                else:
                    img.save(output_path, quality=quality)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True
                })
                logger.info(f"图片格式转换成功: {file_path} -> {output_path}")
            except Exception as e:
                logger.error(f"图片格式转换失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def batch_add_watermark(
        files: List[str],
        watermark_text: str,
        output_dir: Optional[str] = None,
        font_size: int = 24,
        opacity: float = 0.3,
        position: str = "bottom-right"
    ) -> List[Dict[str, Any]]:
        """批量添加文字水印

        Args:
            files: 图片文件列表
            watermark_text: 水印文字
            output_dir: 输出目录，None表示原目录
            font_size: 水印字体大小
            opacity: 水印透明度，0-1
            position: 水印位置，可选值：top-left, top-right, bottom-left, bottom-right, center

        Returns:
            添加水印结果列表
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(
                        output_dir) / f"{file_path.stem}_watermarked{file_path.suffix}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_name(
                        f"{file_path.stem}_watermarked{file_path.suffix}")

                # 打开图片
                img = Image.open(file_path).convert("RGBA")

                # 创建水印图片
                watermark = Image.new("RGBA", img.size, (255, 255, 255, 0))
                draw = ImageDraw.Draw(watermark)

                # 获取字体
                try:
                    font = ImageFont.truetype("arial.ttf", font_size)
                except IOError:
                    font = ImageFont.load_default()

                # 计算水印位置
                text_width, text_height = draw.textsize(
                    watermark_text, font=font)

                if position == "top-left":
                    x, y = 10, 10
                elif position == "top-right":
                    x = img.width - text_width - 10
                    y = 10
                elif position == "bottom-left":
                    x = 10
                    y = img.height - text_height - 10
                elif position == "bottom-right":
                    x = img.width - text_width - 10
                    y = img.height - text_height - 10
                else:  # center
                    x = (img.width - text_width) // 2
                    y = (img.height - text_height) // 2

                # 绘制水印
                draw.text((x, y), watermark_text, font=font,
                          fill=(255, 255, 255, int(255 * opacity)))

                # 合并图片
                result = Image.alpha_composite(img, watermark)
                result = result.convert("RGB")

                # 保存图片
                result.save(output_path)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True
                })
                logger.info(f"图片添加水印成功: {file_path} -> {output_path}")
            except Exception as e:
                logger.error(f"图片添加水印失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def generate_qr_code(
        data: str,
        output_path: str,
        size: int = 300,
        border: int = 4,
        fill_color: str = "black",
        back_color: str = "white"
    ) -> Dict[str, Any]:
        """生成二维码

        Args:
            data: 二维码数据
            output_path: 输出路径
            size: 二维码大小
            border: 边框大小
            fill_color: 填充颜色
            back_color: 背景颜色

        Returns:
            生成结果
        """
        try:
            # 创建二维码对象，提高纠错级别，让库自动选择合适的版本
            qr = qrcode.QRCode(
                version=1,  # 从版本1开始，确保二维码密度适中
                error_correction=qrcode.constants.ERROR_CORRECT_H,  # 最高纠错级别，可纠正30%的错误
                box_size=10,  # 增大box_size，确保二维码元素足够大
                border=border,
            )

            # 添加数据
            qr.add_data(data)
            qr.make(fit=True)

            # 创建图片，使用纯黑白颜色确保对比度足够
            img = qr.make_image(
                fill_color="black",  # 强制使用纯黑色
                back_color="white"   # 强制使用纯白色
            )

            # 不调整大小，保持原始生成的大小，确保二维码元素清晰
            # 直接保存原始大小的二维码，避免缩放导致的模糊
            img.save(output_path, format="PNG", optimize=False, quality=100)

            logger.info(f"二维码生成成功: {output_path}")
            return {
                "success": True,
                "output_path": output_path
            }
        except Exception as e:
            logger.error(f"二维码生成失败: {output_path}, 错误: {e}")
            return {
                "success": False,
                "error": str(e)
            }

    @staticmethod
    def batch_generate_qr_codes(
        data_list: List[Dict[str, str]],
        output_dir: str,
        size: int = 256,
        border: int = 4
    ) -> List[Dict[str, Any]]:
        """批量生成二维码

        Args:
            data_list: 二维码数据列表，每个元素包含data和filename
            output_dir: 输出目录
            size: 二维码大小
            border: 边框大小

        Returns:
            生成结果列表
        """
        results = []

        # 确保输出目录存在
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        for item in data_list:
            try:
                data = item["data"]
                filename = item["filename"]
                output_path = Path(output_dir) / filename

                result = ImageTools.generate_qr_code(
                    data=data,
                    output_path=str(output_path),
                    size=size,
                    border=border
                )

                results.append({
                    "data": data,
                    "filename": filename,
                    "output_path": str(output_path),
                    "success": result["success"]
                })
            except Exception as e:
                logger.error(f"批量二维码生成失败: {item}, 错误: {e}")
                results.append({
                    "data": item.get("data", ""),
                    "filename": item.get("filename", ""),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def batch_apply_filter(
        files: List[str],
        filter_type: str,
        output_dir: Optional[str] = None,
        quality: int = 85
    ) -> List[Dict[str, Any]]:
        """批量应用图片滤镜

        Args:
            files: 图片文件列表
            filter_type: 滤镜类型，可选值：blur, sharpen, grayscale, sepia, contrast, brightness
            output_dir: 输出目录，None表示原目录
            quality: 输出图片质量

        Returns:
            滤镜应用结果列表
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(
                        output_dir) / f"{file_path.stem}_{filter_type}{file_path.suffix}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_name(
                        f"{file_path.stem}_{filter_type}{file_path.suffix}")

                # 打开图片
                img = Image.open(file_path)

                # 应用滤镜
                if filter_type == "blur":
                    result_img = img.filter(ImageFilter.BLUR)
                elif filter_type == "sharpen":
                    result_img = img.filter(ImageFilter.SHARPEN)
                elif filter_type == "grayscale":
                    result_img = ImageOps.grayscale(img)
                elif filter_type == "sepia":
                    # 应用褐色滤镜
                    sepia = img.convert("RGB")
                    width, height = sepia.size
                    pixels = sepia.load()
                    for py in range(height):
                        for px in range(width):
                            r, g, b = pixels[px, py]
                            tr = int(0.393 * r + 0.769 * g + 0.189 * b)
                            tg = int(0.349 * r + 0.686 * g + 0.168 * b)
                            tb = int(0.272 * r + 0.534 * g + 0.131 * b)
                            pixels[px, py] = (
                                min(255, tr), min(255, tg), min(255, tb))
                    result_img = sepia
                elif filter_type == "contrast":
                    enhancer = ImageEnhance.Contrast(img)
                    result_img = enhancer.enhance(1.5)  # 增强对比度
                elif filter_type == "brightness":
                    enhancer = ImageEnhance.Brightness(img)
                    result_img = enhancer.enhance(1.5)  # 增强亮度
                else:
                    result_img = img

                # 保存图片
                result_img.save(output_path, quality=quality)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True,
                    "filter": filter_type
                })
                logger.info(f"图片滤镜应用成功: {file_path} -> {output_path}")
            except Exception as e:
                logger.error(f"图片滤镜应用失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def batch_ocr_images(
        files: List[str],
        language: str = "chi_sim+eng",
        output_dir: Optional[str] = None,
        use_vl_model: bool = False
    ) -> List[Dict[str, Any]]:
        """批量OCR识别图片文字（使用AI，不再依赖本地OCR引擎）

        Args:
            files: 图片文件列表
            language: OCR语言
            output_dir: 输出目录，None表示不保存
            use_vl_model: 是否使用Qwen-VL-Max模型进行OCR

        Returns:
            OCR识别结果列表
        """
        results = []
        
        # 导入API Manager
        from src.modules.api_manager import APIManager
        api_manager = APIManager()

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 调用API Manager的OCR方法，支持使用VL模型
                success, ocr_text, error = api_manager.ocr_image(str(file_path), use_vl_model=use_vl_model)
                
                if not success:
                    raise Exception(error)

                # 保存识别结果
                output_path = None
                if output_dir:
                    output_path = Path(output_dir) / \
                        f"{file_path.stem}_ocr.txt"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                    with open(output_path, "w", encoding="utf-8") as f:
                        f.write(ocr_text)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path) if output_path else None,
                    "success": True,
                    "ocr_text": ocr_text
                })
                logger.info(f"图片OCR识别成功: {file_path}")
            except Exception as e:
                logger.error(f"图片OCR识别失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def image_to_table(
        image_path: str,
        use_vl_model: bool
    ) -> Tuple[bool, Optional[Dict], Optional[str]]:
        """
        将图片中的表格转化为可编辑表格（使用"几何+AI"双驱动模式）

        Args:
            image_path: 图片文件路径
            use_vl_model: 是否使用Qwen-VL-Max模型进行表格识别

        Returns:
            Tuple of (success, table_data, error_message)
        """
        try:
            # 使用新的TableProcessor类进行表格识别
            from src.modules.table_processor import TableProcessor
            
            # 创建TableProcessor实例
            table_processor = TableProcessor()
            
            # 调用TableProcessor的extract_table方法，实现三级降级机制
            success, table_data, error = table_processor.extract_table(image_path)
            
            if success:
                logger.info(f"图片转表格成功: {image_path}")
            else:
                logger.error(f"图片转表格失败: {image_path}, 错误: {error}")
            
            return success, table_data, error
        except Exception as e:
            logger.error(f"图片转表格失败: {image_path}, 错误: {e}")
            return False, None, str(e)
    

    
    @staticmethod
    def export_table(
        table_data: Dict,
        output_path: str,
        export_format: str = "excel"
    ) -> Tuple[bool, Optional[str]]:
        """
        将表格数据导出为Excel或其他格式

        Args:
            table_data: 表格数据字典，包含headers和rows字段，可选merged_cells字段
            output_path: 输出文件路径
            export_format: 导出格式，支持excel、csv等

        Returns:
            Tuple of (success, error_message)
        """
        try:
            import pandas as pd
            
            # 检查表格数据格式
            if not isinstance(table_data, dict) or "rows" not in table_data:
                return False, "无效的表格数据格式，缺少rows字段"
            
            # 获取表格数据
            rows = table_data["rows"]
            
            # 检查是否有数据
            if not rows:
                return False, "表格数据为空"
            
            # 获取表头和合并单元格信息
            headers = table_data.get("headers", [])
            merged_cells = table_data.get("merged_cells", [])
            
            # 确保表头和数据列数匹配
            max_data_cols = max(len(row) for row in rows) if rows else 0
            if headers:
                header_cols = len(headers)
                
                if max_data_cols > header_cols:
                    # 如果数据列数更多，扩展表头
                    headers.extend(['' for _ in range(header_cols, max_data_cols)])
                elif max_data_cols < header_cols:
                    # 如果数据列数更少，截断表头
                    headers = headers[:max_data_cols]
            
            # 根据导出格式保存文件
            if export_format.lower() == "excel":
                # 导出为Excel文件，支持合并单元格
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Border, Side
                
                # 定义单元格样式函数
                def apply_style(cell, is_header=False):
                    """应用单元格样式"""
                    # 居中对齐能极大程度在感官上减少“错位感”
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    # 加上细边框，视觉上更接近原图
                    thin = Side(border_style="thin", color="000000")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                
                # 创建Workbook对象
                wb = Workbook()
                ws = wb.active
                
                # 计算总行数和总列数
                total_rows = len(rows) + 1  # 1行表头
                total_cols = max(max_data_cols, len(headers)) if headers else max_data_cols
                
                # 初始化布尔占用矩阵
                # occupancy_matrix[r][c] = True 表示该单元格已被合并单元格占用
                occupancy_matrix = [[False for _ in range(total_cols)] for _ in range(total_rows)]
                
                # 写入表头并应用样式
                if headers:
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        apply_style(cell, is_header=True)
                
                # 写入数据行并应用样式
                data_row_start = 2  # 数据行从第2行开始
                for row_idx, row_data in enumerate(rows, 1):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        # Excel实际行号
                        excel_row = data_row_start - 1 + row_idx
                        excel_col = col_idx
                        
                        # 跳过已被占用的单元格
                        if occupancy_matrix[excel_row - 1][excel_col - 1]:
                            continue
                        
                        cell = ws.cell(row=excel_row, column=excel_col, value=cell_value)
                        apply_style(cell, is_header=False)
                
                # 处理合并单元格
                if merged_cells:
                    print("  开始处理合并单元格...")
                    
                    # 应用合并单元格
                    for merge_info in merged_cells:
                        if len(merge_info) == 4:
                            start_row, start_col, end_row, end_col = merge_info
                            
                            # 转换为Excel行号（Excel行号从1开始）
                            excel_start_row = start_row + 1
                            excel_end_row = end_row + 1
                            excel_start_col = start_col + 1
                            excel_end_col = end_col + 1
                            
                            # 确保范围有效
                            if (excel_start_row > total_rows or excel_start_col > total_cols or
                                excel_end_row < 1 or excel_end_col < 1):
                                continue
                            
                            # 限制合并范围在有效范围内
                            excel_end_row = min(excel_end_row, total_rows)
                            excel_end_col = min(excel_end_col, total_cols)
                            
                            # 只合并有效范围的单元格（行或列跨度>1）
                            if excel_end_row > excel_start_row or excel_end_col > excel_start_col:
                                try:
                                    print(f"  应用合并单元格: 行{excel_start_row}-{excel_end_row}, 列{excel_start_col}-{excel_end_col}")
                                    ws.merge_cells(start_row=excel_start_row, start_column=excel_start_col,
                                                end_row=excel_end_row, end_column=excel_end_col)
                                    
                                    # 设置合并单元格的样式
                                    merged_cell = ws.cell(row=excel_start_row, column=excel_start_col)
                                    apply_style(merged_cell, is_header=(excel_start_row == 1))
                                    
                                    # 更新布尔占用矩阵
                                    for r in range(excel_start_row - 1, excel_end_row):
                                        for c in range(excel_start_col - 1, excel_end_col):
                                            occupancy_matrix[r][c] = True
                                except Exception as e:
                                    print(f"  ⚠️  合并单元格失败: {e}, 范围: ({excel_start_row}, {excel_start_col}, {excel_end_row}, {excel_end_col})")
                
                # 确保文件不存在，使用更严格的文件处理
            import time
            import tempfile
            import shutil
            import uuid
            
            # 使用临时文件策略避免文件锁定问题
            temp_path = None
            success = False
            
            # 尝试多次写入，解决文件占用问题
            for attempt in range(3):
                try:
                    # 创建唯一的临时文件路径
                    temp_dir = tempfile.gettempdir()
                    temp_filename = f"temp_excel_{uuid.uuid4().hex}.xlsx"
                    temp_path = os.path.join(temp_dir, temp_filename)
                    
                    # 保存到临时文件
                    wb.save(temp_path)
                    
                    # 确保输出目录存在
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    
                    # 尝试删除目标文件（如果存在）
                    if os.path.exists(output_path):
                        os.remove(output_path)
                    
                    # 使用shutil.move原子操作替换目标文件
                    shutil.move(temp_path, output_path)
                    success = True
                    temp_path = None  # 标记临时文件已处理
                    break
                except PermissionError:
                    # 文件被占用，等待后重试
                    time.sleep(1)
                except Exception as e:
                    # 其他错误，跳出循环
                    print(f"  ⚠️  Excel导出重试失败: {e}")
                    break
                finally:
                    # 清理临时文件（如果有）
                    if temp_path and os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                        except:
                            pass
            
            if not success:
                # 最终尝试直接保存
                try:
                    # 生成一个新的文件名
                    base_name, ext = os.path.splitext(output_path)
                    backup_path = f"{base_name}_{uuid.uuid4().hex[:8]}{ext}"
                    print(f"  ⚠️  无法覆盖现有文件，保存为备份文件: {backup_path}")
                    wb.save(backup_path)
                    # 更新输出路径为备份路径
                    output_path = backup_path
                except Exception as e:
                    raise Exception(f"Excel导出失败: {e}")
            elif export_format.lower() == "csv":
                # CSV格式不支持合并单元格，使用pandas导出
                try:
                    # 创建DataFrame
                    if headers:
                        df = pd.DataFrame(rows, columns=headers)
                    else:
                        df = pd.DataFrame(rows)
                except Exception as e:
                    print(f"  ⚠️  创建DataFrame失败，使用备用方案: {e}")
                    # 备用方案：直接使用数据创建，不指定表头
                    df = pd.DataFrame(rows)
                
                # 替换所有NaN值为空字符串
                df = df.fillna('')
                
                # 保存CSV文件
                if os.path.exists(output_path):
                    os.remove(output_path)
                df.to_csv(output_path, index=False, encoding="utf-8-sig")
            else:
                return False, f"不支持的导出格式: {export_format}"
            
            logger.info(f"表格导出成功: {output_path}")
            return True, None
        except ImportError as e:
            logger.error(f"导出表格失败: 缺少必要的库 - {e}")
            return False, f"缺少必要的库: {e}"
        except Exception as e:
            logger.error(f"导出表格失败: {e}")
            return False, str(e)
    
    @staticmethod
    def batch_crop_images(
        files: List[str],
        left: int = 0,
        top: int = 0,
        right: Optional[int] = None,
        bottom: Optional[int] = None,
        output_dir: Optional[str] = None,
        quality: int = 85
    ) -> List[Dict[str, Any]]:
        """批量裁剪图片

        Args:
            files: 图片文件列表
            left: 左裁剪位置
            top: 上裁剪位置
            right: 右裁剪位置，None表示图片宽度
            bottom: 下裁剪位置，None表示图片高度
            output_dir: 输出目录，None表示原目录
            quality: 输出图片质量

        Returns:
            裁剪结果列表
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(
                        output_dir) / f"{file_path.stem}_cropped{file_path.suffix}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_name(
                        f"{file_path.stem}_cropped{file_path.suffix}")

                # 打开图片
                img = Image.open(file_path)

                # 计算裁剪区域
                img_width, img_height = img.size

                crop_right = right or img_width
                crop_bottom = bottom or img_height

                # 执行裁剪
                cropped_img = img.crop((left, top, crop_right, crop_bottom))

                # 保存图片
                cropped_img.save(output_path, quality=quality)

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True,
                    "crop_region": (left, top, crop_right, crop_bottom)
                })
                logger.info(f"图片裁剪成功: {file_path} -> {output_path}")
            except Exception as e:
                logger.error(f"图片裁剪失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results

    @staticmethod
    def merge_images(
        files: List[str],
        output_path: str,
        layout: str = "horizontal",
        quality: int = 85
    ) -> Dict[str, Any]:
        """合并多张图片

        Args:
            files: 图片文件列表
            output_path: 输出路径
            layout: 合并布局，horizontal或vertical
            quality: 输出图片质量

        Returns:
            合并结果
        """
        try:
            # 打开所有图片
            images = [Image.open(file_path) for file_path in files]

            # 计算合并后的尺寸
            if layout == "horizontal":
                total_width = sum(img.width for img in images)
                max_height = max(img.height for img in images)
                result_img = Image.new(
                    "RGB", (total_width, max_height), (255, 255, 255))

                # 合并图片
                x_offset = 0
                for img in images:
                    y_offset = (max_height - img.height) // 2
                    result_img.paste(img, (x_offset, y_offset))
                    x_offset += img.width
            else:  # vertical
                max_width = max(img.width for img in images)
                total_height = sum(img.height for img in images)
                result_img = Image.new(
                    "RGB", (max_width, total_height), (255, 255, 255))

                # 合并图片
                y_offset = 0
                for img in images:
                    x_offset = (max_width - img.width) // 2
                    result_img.paste(img, (x_offset, y_offset))
                    y_offset += img.height

            # 保存图片
            result_img.save(output_path, quality=quality)

            logger.info(f"图片合并成功: {files} -> {output_path}")
            return {
                "success": True,
                "output_path": output_path,
                "merged_files": files
            }
        except Exception as e:
            logger.error(f"图片合并失败: {output_path}, 错误: {e}")
            return {
                "success": False,
                "error": str(e)
            }

    @staticmethod
    def crop_to_content(img: np.ndarray) -> np.ndarray:
        """
        裁剪图片到内容区域，去除周围的白边
        
        Args:
            img: 输入图片
            
        Returns:
            np.ndarray: 裁剪后的图片
        """
        # 转换为灰度图
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # 二值化处理
        _, binary = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY_INV)
        
        # 查找轮廓
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if not contours:
            return img
        
        # 找到最大的轮廓
        max_contour = max(contours, key=cv2.contourArea)
        
        # 计算轮廓的边界矩形
        x, y, w, h = cv2.boundingRect(max_contour)
        
        # 裁剪图片
        cropped = img[y:y+h, x:x+w]
        
        return cropped
    
    @staticmethod
    def batch_compress_images(
        files: List[str],
        max_size: int = 1024,
        quality: int = 85,
        output_dir: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """批量压缩图片

        Args:
            files: 图片文件列表
            max_size: 最大尺寸（宽或高）
            quality: 输出图片质量
            output_dir: 输出目录，None表示原目录

        Returns:
            压缩结果列表
        """
        results = []

        for file_path in files:
            try:
                file_path = Path(file_path)
                if not file_path.exists() or not file_path.is_file():
                    continue

                # 确定输出路径
                if output_dir:
                    output_path = Path(
                        output_dir) / f"{file_path.stem}_compressed{file_path.suffix}"
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                else:
                    output_path = file_path.with_name(
                        f"{file_path.stem}_compressed{file_path.suffix}")

                # 打开图片
                img = Image.open(file_path)

                # 计算缩放比例
                img_width, img_height = img.size
                if img_width > max_size or img_height > max_size:
                    if img_width > img_height:
                        ratio = max_size / img_width
                    else:
                        ratio = max_size / img_height

                    new_width = int(img_width * ratio)
                    new_height = int(img_height * ratio)

                    # 调整大小
                    img = img.resize((new_width, new_height), Image.LANCZOS)

                # 保存图片
                img.save(output_path, quality=quality)

                # 计算压缩率
                original_size = os.path.getsize(file_path)
                compressed_size = os.path.getsize(output_path)
                compression_ratio = (1 - compressed_size / original_size) * 100

                results.append({
                    "file_path": str(file_path),
                    "output_path": str(output_path),
                    "success": True,
                    "original_size": original_size,
                    "compressed_size": compressed_size,
                    "compression_ratio": compression_ratio
                })
                logger.info(
                    f"图片压缩成功: {file_path} -> {output_path}, 压缩率: {compression_ratio:.2f}%")
            except Exception as e:
                logger.error(f"图片压缩失败: {file_path}, 错误: {e}")
                results.append({
                    "file_path": str(file_path),
                    "success": False,
                    "error": str(e)
                })

        return results
